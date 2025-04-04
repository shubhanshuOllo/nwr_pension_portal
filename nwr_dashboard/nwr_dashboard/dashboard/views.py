from django.shortcuts import render,redirect
from django.http import JsonResponse,HttpResponse
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from .models import *
import datetime
from django.db import transaction
from django.contrib.auth.decorators import login_required
from django.db import connection, transaction
import pandas as pd
import datetime
import io
from .rules import *
from django.db.models import Q
from .rules import *
import re
import openai
import json
from decimal import Decimal
import os
from dotenv import load_dotenv
from io import BytesIO
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import openpyxl
from datetime import datetime

load_dotenv()


@csrf_exempt
def dashboard(request):
    if not request.user.is_authenticated:
        return redirect("login") 
    return render(request, "dashboard.html")



def bulk_update_records(update_records):
    """
    Perform bulk updates using raw SQL.
    """
    update_query = """
        UPDATE nwr_master_data
        SET 
            ppo_number = %(ppo_number)s,
            name = %(name)s,
            dob = %(dob)s,
            pension_start_date = %(pension_start_date)s,
            date_of_retirement = %(date_of_retirement)s,
            age = %(age)s
        WHERE id = %(id)s
    """
    
    with connection.cursor() as cursor:
        cursor.executemany(update_query, update_records)


def bulk_insert_records(new_records):
    """
    Perform a bulk insert using raw SQL (COPY for PostgreSQL or executemany for other databases).
    """
    if not new_records:
        return

    with connection.cursor() as cursor:
        if connection.vendor == 'postgresql':
            # Use COPY for PostgreSQL (much faster than INSERT)
            output = io.StringIO()
            for record in new_records:
                output.write(
                    f"{record['account_number']},{record['ppo_number']},{record['name']},{record['dob']},{record['pension_start_date']},{record['date_of_retirement']},{record['age']}\n"
                )
            output.seek(0)
            cursor.copy_from(output, 'nwr_master_data', sep=',', columns=[
                'account_number', 'ppo_number', 'name', 'dob', 'pension_start_date', 'date_of_retirement', 'age'
            ])
        else:
            # Use executemany() for MySQL or SQLite
            insert_query = """
                INSERT INTO nwr_master_data 
                (account_number, ppo_number, name, dob, pension_start_date, date_of_retirement, age)
                VALUES (%(account_number)s, %(ppo_number)s, %(name)s, %(dob)s, %(pension_start_date)s, %(date_of_retirement)s, %(age)s)
            """
            cursor.executemany(insert_query, new_records)
@csrf_exempt                                                                          
@login_required             
def upload_master_excel(request):            
    if request.method == 'POST':              
        if 'file' in request.FILES:
            excel_file = request.FILES['file']

            
            
            df = pd.read_excel(excel_file, skiprows=1)
            
            
            df = df.drop_duplicates(subset=['ACCOUNT_NUMBER'])

            
            date_cols = ['DATE_OF_BIRTH', 'DATE_START', 'DATE_OF_RETIREMENT']
            for col in date_cols:
                df[col] = pd.to_datetime(df[col], errors='coerce')

            today = datetime.date.today()

            
            existing_accounts = {}
            with connection.cursor() as cursor:
                cursor.execute("SELECT id, account_number FROM nwr_master_data")
                for record in cursor.fetchall():
                    acc_num = str(record[1]).strip()
                    existing_accounts.setdefault(acc_num, []).append(record[0])  

            new_records = []
            update_records = []

            for _, row in df.iterrows():
                try:
                    acc_num = str(row['ACCOUNT_NUMBER']).strip() if pd.notna(row['ACCOUNT_NUMBER']) else None
                    if not acc_num:
                           continue  

                    dob = row['DATE_OF_BIRTH'].date() if pd.notna(row['DATE_OF_BIRTH']) else None
                    age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day)) if dob else 0

                    record_data = {
                        "account_number": acc_num,
                        "ppo_number": str(row['PPO_NUMBER']).strip() if pd.notna(row['PPO_NUMBER']) else None,
                        "name": str(row['NAME']).strip() if pd.notna(row['NAME']) else None,
                        "dob": dob,
                        "pension_start_date": row['DATE_START'].date() if pd.notna(row['DATE_START']) else None,
                        "date_of_retirement": row['DATE_OF_RETIREMENT'].date() if pd.notna(row['DATE_OF_RETIREMENT']) else None,
                        "age": age
                    }

                    if acc_num in existing_accounts:
                       
                        for record_id in existing_accounts[acc_num]:
                            record_data["id"] = record_id  
                            update_records.append(record_data)
                    else:
                        
                        new_records.append(record_data)

                except Exception as e:
                    print(f" Error processing row {row}: {e}")  
        
           
            if new_records:
                bulk_insert_records(new_records)

            
            chunk_size = 500  
            if update_records:
                with transaction.atomic():
                    for i in range(0, len(update_records), chunk_size):
                        bulk_update_records(update_records[i:i+chunk_size])

            print(" Upload completed successfully!")  # Debugging
            return JsonResponse({'status': 'success'}, status=200)
        
        return JsonResponse({'status': 'failed', 'message': 'No file provided'}, status=400)
    
    return JsonResponse({'status': 'failed', 'message': 'Invalid request method'}, status=400)


# @csrf_exempt
# def bulk_update_nwr_zone(update_records):
#     """
#     Perform bulk updates using raw SQL.
#     """
#     update_query = """
#         UPDATE nwr_zone_data
#         SET 
#             ppo_zone_code = %(ppo_zone_code)s,
#             pensioner_id = %(pensioner_id)s,
#             old_ppo = %(old_ppo)s,
#             new_ppo = %(new_ppo)s,
#             emp_name = %(emp_name)s,
#             gender = %(gender)s,
#             cessation_date = %(cessation_date)s,
#             death_date = %(death_date)s,
#             pension_amount = %(pension_amount)s,
#             efp_amount = %(efp_amount)s,
#             efp_date = %(efp_date)s,
#             ppb_account_number = %(ppb_account_number)s
#         WHERE id = %(id)s
#     """
    
#     with connection.cursor() as cursor:
#         cursor.executemany(update_query, update_records)


# def bulk_insert_nwr_zone(new_records):
#     """
#     Perform bulk insert using raw SQL (COPY for PostgreSQL or executemany for other databases).
#     """
#     if not new_records:
#         return

#     with connection.cursor() as cursor:
#         if connection.vendor == 'postgresql':
#             # Use COPY for PostgreSQL (much faster than INSERT)
#             output = io.StringIO()
#             for record in new_records:
#                 output.write(
#                     f"{record['ppo_zone_code']},{record['pensioner_id']},{record['old_ppo']},{record['new_ppo']},{record['emp_name']},{record['gender']},{record['cessation_date']},{record['death_date']},{record['pension_amount']},{record['efp_amount']},{record['efp_date']},{record['ppb_account_number']}\n"
#                 )
#             output.seek(0)
#             cursor.copy_from(output, 'nwr_zone_data', sep=',', columns=[
#                 'ppo_zone_code', 'pensioner_id', 'old_ppo', 'new_ppo', 'emp_name', 'gender', 'cessation_date',
#                 'death_date', 'pension_amount', 'efp_amount', 'efp_date', 'ppb_account_number'
#             ])
#         else:
#             # Use executemany() for MySQL or SQLite
#             insert_query = """
#                 INSERT INTO nwr_zone_data 
#                 (ppo_zone_code, pensioner_id, old_ppo, new_ppo, emp_name, gender, cessation_date, death_date, pension_amount, efp_amount, efp_date, ppb_account_number)
#                 VALUES (%(ppo_zone_code)s, %(pensioner_id)s, %(old_ppo)s, %(new_ppo)s, %(emp_name)s, %(gender)s, %(cessation_date)s, %(death_date)s, %(pension_amount)s, %(efp_amount)s, %(efp_date)s, %(ppb_account_number)s)
#             """
#             cursor.executemany(insert_query, new_records)


@login_required
@csrf_exempt
def upload_nwr_zone(request):
    if request.method == 'POST':
        if 'file' in request.FILES:
            excel_file = request.FILES['file']
            
            
            df = pd.read_excel(excel_file)
            
            
            df = df.drop_duplicates(subset=['PPO_NUMBER'])

            # Convert date columns safely
            date_cols = ['CESSATION_DATE', 'DEATH_DATE', 'EFP_DATE']
            for col in date_cols:
                df[col] = pd.to_datetime(df[col], errors='coerce')

            new_records = []
            update_records = []

          
            existing_records = {}
            with connection.cursor() as cursor:
                cursor.execute("SELECT id, new_ppo FROM nwr_zone_data")
                for record in cursor.fetchall():
                    existing_records[str(record[1]).strip()] = record[0]

            for _, row in df.iterrows():
                try:
                    new_ppo = str(row['PPO_NUMBER']).strip() if pd.notna(row['PPO_NUMBER']) else None
                    if not new_ppo:
                        continue  

                    record_data = {
                        "ppo_zone_code": str(row['PPO_ZONE_CODE']).strip(),
                        "pensioner_id": row['PENSIONER_ID'],
                        "old_ppo": str(row['OLD_PPO_NUMBER']).strip(),
                        "new_ppo": new_ppo,
                        "emp_name": str(row['EMP_NAME']).strip(),
                        "gender": str(row['GENDER']).strip(),
                        "cessation_date": row['CESSATION_DATE'].date() if pd.notna(row['CESSATION_DATE']) else None,
                        "death_date": row['DEATH_DATE'].date() if pd.notna(row['DEATH_DATE']) else None,
                        "pension_amount": row['PENSION_AMOUNT'] if pd.notna(row['PENSION_AMOUNT']) else 0,
                        "efp_amount": row['EFP_AMOUNT'] if pd.notna(row['EFP_AMOUNT']) else 0,
                        "efp_date": row['EFP_DATE'].date() if pd.notna(row['EFP_DATE']) else None,
                        "ppb_account_number": clean_account_number(row['PPB_ACCOUNT_NUMBER'])
                    }

                    existing_id = existing_records.get(new_ppo)  
                    if existing_id:
                        record_data["id"] = existing_id
                        update_records.append(record_data)
                    else:
                        new_records.append(record_data)

                except Exception as e:
                    print(f" Error processing row {row}: {e}")  

            #  Reduce batch size for updates
            chunk_size = 200  

            #  Bulk Insert New Records
            if new_records:
                with connection.cursor() as cursor:
                    insert_query = """
                        INSERT INTO nwr_zone_data 
                        (ppo_zone_code, pensioner_id, old_ppo, new_ppo, emp_name, gender, cessation_date, death_date, pension_amount, efp_amount, efp_date, ppb_account_number)
                        VALUES (%(ppo_zone_code)s, %(pensioner_id)s, %(old_ppo)s, %(new_ppo)s, %(emp_name)s, %(gender)s, %(cessation_date)s, %(death_date)s, %(pension_amount)s, %(efp_amount)s, %(efp_date)s, %(ppb_account_number)s)
                    """
                    cursor.executemany(insert_query, new_records)

            #  Bulk Update Existing Records (Using Transactions for Faster Commits)
            if update_records:
                update_query = """
                    UPDATE nwr_zone_data
                    SET 
                        ppo_zone_code = %(ppo_zone_code)s,
                        pensioner_id = %(pensioner_id)s,
                        old_ppo = %(old_ppo)s,
                        emp_name = %(emp_name)s,
                        gender = %(gender)s,
                        cessation_date = %(cessation_date)s,
                        death_date = %(death_date)s,
                        pension_amount = %(pension_amount)s,
                        efp_amount = %(efp_amount)s,
                        efp_date = %(efp_date)s,
                        ppb_account_number = %(ppb_account_number)s
                    WHERE id = %(id)s
                """
                with transaction.atomic():
                    for i in range(0, len(update_records), chunk_size):
                        with connection.cursor() as cursor:
                            cursor.executemany(update_query, update_records[i:i+chunk_size])

            print(" Upload completed successfully!")  
            return JsonResponse({'status': 'success'}, status=200)
        
        return JsonResponse({'status': 'failed', 'message': 'No file provided'}, status=400)
    
    return JsonResponse({'status': 'failed', 'message': 'Invalid request method'}, status=400)


@login_required
@csrf_exempt
def debit_scroll(request):
    if request.method == 'POST':
        if 'file' in request.FILES:
            excel_file = request.FILES['file']
            xls = pd.ExcelFile(excel_file)
            selected_sheet = max(xls.sheet_names, key=lambda sheet: len(pd.read_excel(excel_file, sheet_name=sheet)))

            # Read the sheet with the most rows
            df = pd.read_excel(excel_file, sheet_name=selected_sheet)
            print(df)
 
            try:
                df = df.astype(str)
                column_indexes = [0, 3, 5, 6, 7, 9, 10, 11, 14, 19, 26,29]
                df = df.iloc[:, column_indexes]
                df.columns = [
                    'file_number', 'type_of_pension', 'new_ppo', 'old_ppo', 'current_pensioner',
                    'pension_month', 'basic_pension', 'deduction', 'fma', 'da', 'pension','account_number'
                ]
               
                df['new_ppo'] = df['new_ppo'].replace('nan', '')
                df = df.replace('nan', '')
 
                numeric_columns = [5, 6, 7, 8, 9, 10]
                for col in numeric_columns:
                    df.iloc[:, col] = pd.to_numeric(df.iloc[:, col], errors='coerce').fillna(0).astype(int)
 
                # print(df.head())
 
               
                # # sys.exit("Debugging exit point")
 
            except Exception as e:
                print("Error processing data:", e)
                return JsonResponse({"error": str(e)}, status=400)
 
           
            
            try:
                records = df.to_dict(orient='records')
                batch_size = 200  # Prevents database overload
 
                for i in range(0, len(records), batch_size):
                    batch = records[i:i+batch_size]
 
                    # Extract unique keys for existing records
                    lookup_keys = [
                        (row.get('old_ppo', ''), row.get('new_ppo', ''), int(float(row.get('pension_month', 0))))
                        for row in batch
                    ]
 
                    # Fetch existing records from DB
                    existing_records = DebitScroll.objects.filter(
                        Q(old_ppo__in=[key[0] for key in lookup_keys]) &
                        Q(new_ppo__in=[key[1] for key in lookup_keys]) &
                        Q(pension_month__in=[key[2] for key in lookup_keys])
                    ).values_list('old_ppo', 'new_ppo', 'pension_month', 'id')
 
                    existing_dict = {(x[0], x[1], x[2]): x[3] for x in existing_records}
 
                    to_update = []
                    to_insert = []
 
                    for row in batch:
                        key = (row.get('old_ppo', ''), row.get('new_ppo', ''), int(float(row.get('pension_month', 0))))
                        acc_num = clean_account_number(row.get('account_number', ''))
 
                        if key in existing_dict:
                            # Update existing record
                            record = DebitScroll(
                                id=existing_dict[key],  # Use existing ID to update
                                file_number=row.get('file_number', ''),
                                type_of_pension=row.get('type_of_pension', None),
                                current_pensioner=row.get('current_pensioner', ''),
                                basic_pension=int(float(row.get('basic_pension', 0))),
                                deduction=int(float(row.get('deduction', 0))),
                                fma=int(float(row.get('fma', 0))),
                                da=int(float(row.get('da', 0))),
                                pension=int(float(row.get('pension', 0))),
                                account_number=acc_num  # Default to an empty string if missing

                            )
                            to_update.append(record)
                        else:
                            acc_num = clean_account_number(row.get('account_number', ''))
                            # Insert new record
                            to_insert.append(DebitScroll(
                                file_number=row.get('file_number', ''),
                                type_of_pension=row.get('type_of_pension', None),
                                old_ppo=row.get('old_ppo', ''),
                                new_ppo=row.get('new_ppo', ''),
                                current_pensioner=row.get('current_pensioner', ''),
                                pension_month=int(float(row.get('pension_month', 0))),
                                basic_pension=int(float(row.get('basic_pension', 0))),
                                deduction=int(float(row.get('deduction', 0))),
                                fma=int(float(row.get('fma', 0))),
                                da=int(float(row.get('da', 0))),
                                pension=int(float(row.get('pension', 0))),
                                account_number=acc_num
                            ))
 
                    with transaction.atomic():
                        # Perform bulk update
                        if to_update:
                            DebitScroll.objects.bulk_update(to_update, [
                                'file_number', 'type_of_pension', 'current_pensioner', 'basic_pension',
                                'deduction', 'fma', 'da', 'pension','account_number'
                            ])
                           
 
                        # Perform bulk insert
                        if to_insert:
                            DebitScroll.objects.bulk_create(to_insert)
                return JsonResponse({'status': 'success'}, status=200)        

            except Exception as e:
                print(" Critical Error:", e)    
 
def clean_account_number(value):
    """Convert to string, strip spaces, and remove decimals if any."""
    if pd.isna(value) or value in [None, "nan", "NaN"]:  
        return None  # Handle NaN properly
    return str(value).split(".")[0].strip()



@csrf_exempt
# @login_required
def get_rule(request):
    # Extract month and rule from request
    # month = request.GET.get('month')
    month = request.GET.getlist("month")  
    print(request.GET.keys(),"request")
    rule = request.GET.get("rule", "").strip()
    print(month,"month")
    print(rule,"rule")
    # generate_data_excel()
    
    if len(month)> 1:
        if month[0] == month[1]:
            month_numbers = [int(m.split("-")[1]) for m in month if "-" in m]
            month_numbers.pop(1) 
            month = month_numbers 
        else:
            month_numbers = [int(m.split("-")[1]) for m in month if "-" in m]
            month = month_numbers
    else:
        month_numbers = [int(m.split("-")[1]) for m in month if "-" in m]
        month = month_numbers
        # else:



    
    # month = data.get(month)
    # rule = request.GET.get('rule', '').strip()

   
    data = []
    rule_comp = 0
    
    if len(month)> 1 and rule in ["1","2","3","6"]:
        
        rule_comp = 1
        if rule == "1":                                                                                                                 
            data = comp_overall_payment(month)
        if rule == "2":                                                                                                                                
            data = comp_age_metrics(month)                                      
        if rule == "3":                                                                                                                                                                          
            data = comp_family_pension_conversion(month)
        if rule == "4":                                                                 
            data = comp_revised_pensioners(month)
        if rule == "6":
            data=comp_active_pensioner(month)                                                                                                   
        
    else:
        month_temp = month[0]
        month = month_temp
        if rule == "1":  
            data = overall_payment(month)

        if rule == "2":  
            data = age_metrics(month)

        if rule == "3":  
            data = family_pension_conversion(month)

        if rule == "4":  
            data = revised_pensioners(month)

        if rule == "5":
            data=get_pension_stats_last_6_months(month)

        if rule == "6":
            data=active_pensioner(month)

        if rule == "7":
            data=get_efp_count()

        if rule == "8":
            data = comp_age_bracket()
        if rule == "9":
            data = comp_stopped_pensioner()
        if rule == "10":
            data = commutation_data(month)
    data["rule_comp"] = rule_comp
    data["rule"] = rule
    print(data,"data")
    
    return JsonResponse(data, safe=False)






def upload_mismatch(request):
    if request.method == 'POST':
        if 'file' in request.FILES:
            excel_file = request.FILES['file']
            file_name = excel_file.name.lower()
            if file_name.endswith('.xlsx'):                                             
                engine = "openpyxl"  # For modern Excel files
            elif file_name.endswith('.xls'):
                engine = "xlrd"  # For older Excel files
            
            
            
            try:
                # Read Excel file into a DataFrame
                df = pd.read_excel(excel_file, engine=engine,skiprows=4, header=None)  # Read without headers
                df = df.iloc[:, 1:] 

                # Extract value from the fourth row (index 3)
                # fourth_row_text = str(df.iloc[3, 0])  # Get first column of fourth row

                # # Use regex to extract the last numeric value (202409)
                # match = re.search(r"(\d+)$", fourth_row_text)
                # month = match.group(1) if match else None

                # print("Extracted Value:", month)  # Debugging
                
                # # Start DataFrame from the fifth row (index 4)
                # df = df.iloc[5:].reset_index(drop=True)

                # # Set proper column names (assuming your Excel file has correct headers)
                # df.columns = ["ARPAN PPO Number", "IFSC Code", "Scroll Pension Type", "ARPAN Pension Type", 
                #               "PPO Number", "Pensioner Name", "Scroll Acc No", "Cessation Date", 
                #               "ARPAN Basic", "Scroll Basic", "Basic Diff"]

                df.columns = ["debit_zone_code", "debit_zone", "bank_code", "scroll_ppo_no", 
                          "account_number", "pensioner_name", "ifsc_code"]

              
                # df["Cessation Date"] = pd.to_datetime(df["Cessation Date"], format="%d-%m-%Y", errors="coerce")

                # # Convert datetime to only date (removes time component)
                # df["Cessation Date"] = df["Cessation Date"].apply(lambda x: x.date() if pd.notna(x) else None)


                # Save data to the database
                # for _, row in df.iterrows():
                #     mismatch_data.objects.create(
                #         arpan_ppo_number=row["ARPAN PPO Number"],
                #         ifsc_code=row["IFSC Code"],
                #         scroll_pension_type=row["Scroll Pension Type"],
                #         arpan_pension_type=row["ARPAN Pension Type"],
                #         ppo_number=row["PPO Number"],
                #         pensioner_name=row["Pensioner Name"],
                #         scroll_acc_no=row["Scroll Acc No"],
                #         cessation_date=row["Cessation Date"],
                #         arpan_basic=row["ARPAN Basic"],
                #         scroll_basic=row["Scroll Basic"],
                #         basic_diff=row["Basic Diff"],
                #         month = month
                #     )
                records = [
                arpan_exception(
                    debit_zone_code=row["debit_zone_code"],
                    debit_zone=row["debit_zone"],
                    bank_code=row["bank_code"],
                    scroll_ppo_no=row["scroll_ppo_no"],
                    account_number=row["account_number"],
                    pensioner_name=row["pensioner_name"],
                    ifsc_code=row["ifsc_code"]
                ) 
                for _, row in df.iterrows()
            ]
            
            # Bulk insert data into the database
                arpan_exception.objects.bulk_create(records)

                return JsonResponse({"message": "Data uploaded successfully!"}, status=201)

                # return JsonResponse({'status': 'success'}, status=200)  

            except Exception as e:
                print("Critical Error:", e)  


def process_commutational_data():
    file = r'files\CommutationMismatchData (27) - Oct 24.xls'
    # files\CommutationMismatchData (28) - Sept 24 (2).xls
    try:
        df = pd.read_excel(file, skiprows=4, engine='xlrd')
    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        return

    # Convert Cessation Date column to datetime format
    df['Cessation Date'] = pd.to_datetime(df['Cessation Date'], format="%d-%m-%Y", errors='coerce')

    records = []
    
    for _, row in df.iterrows():
        cessation_date = row['Cessation Date']
        if pd.isna(cessation_date):  # Handle NaT values
            cessation_date = None

        records.append(Commutational_data(
            arpan_ppo_number=row.get('ARPAN PPO Number', None),
            ifsc_code=row.get('IFSC Code', None),
            bank_ppo_number=row.get('Bank PPO Number', None),
            pensioner_name=row.get('Pensioner Name', None),
            scroll_acc_no=row.get('Scroll Acc No', None),
            cessation_date=cessation_date,
            arpan_basic=row.get('ARPAN Basic', None),
            scroll_basic=row.get('Scroll Basic', None),
            basic_diff=row.get('Basic Diff', None),
            arpan_commutation=row.get('ARPAN Commutation', None),
            scroll_commutation=row.get('Scroll Commutation', None),
            commutation_diff=row.get('Commutation Diff', None),
            gen_pdf=row.get('GEN_PDF', None),
            month='202410'
        ))
    print(records)
    if records:
        Commutational_data.objects.bulk_create(records)
        print(f"Successfully inserted {len(records)} records.")
    else:
        print("No records found in the Excel file.")

def load_mismatch_page(request):
    """Load the mismatch.html template."""
    return render(request, 'mismatch.html')

def is_safe_sql(query):
    forbidden_keywords = ["INSERT", "UPDATE", "DELETE", "DROP", "ALTER", "TRUNCATE", "CREATE", "REPLACE"]
    
    # Remove comments to prevent bypass
    query = re.sub(r"--.*?\n|/\*.*?\*/", "", query, flags=re.S)
    
    return not any(re.search(rf"\b{keyword}\b", query, re.IGNORECASE) for keyword in forbidden_keywords)




db_prompt = """You are an expert SQL assistant for a pension management system. The database contains tables related to pensioners, debit transactions, pension mismatches, and zone data. You understand the schema and generate only safe, read-only SQL queries based on user requests. You never generate INSERT, UPDATE, DELETE, or DROP queries. The year is 2024 whenever someone does not mention the year in a query assume 2024.


Here is the database schema:

Tables and Relationships:
Type of pension if f for family and r for regualar pension. If you use debit scroll for writing any query and the user has not provided any month then use 202409 as month.When using debit scroll for any query always set month as constraint.

arpan_exception: It stores data of unlinked pensioners. Unlinked pensioners are those who are present in debit scroll but not present in nwr_zone data. So any query regarding unlinked pensioner shall be answered from this. Stores pensioner exceptions with fields like debit_zone_code, bank_code, scroll_ppo_no, account_number, and ifsc_code,pensioner_name.
debit_scroll: Contains transaction records, including file_number, type_of_pension, new_ppo, current_pensioner, pension_month, and financial details (basic_pension, deduction, da, etc.),pension.For any query regarding pension or overlay use the pension field.

mismatch_data: It stores data which has different values in nwr_zone data and debit scroll. So any query with mismatch shall be answered from this. Identifies pension mismatches based on arpan_ppo_number, ifsc_code, arpan_pension_type, scroll_pension_type, and basic_diff, month. Overpayment and underment shall be calculated based on basic_diff. 

commutational_data: This table stores mismatches in commutation values between nwr_zone_data and debit_scroll. If a query is specifically about commutation mismatch, fetch data from this table. Overpayment and underpayment should be calculated based on commutation_diff we have month colummn in it with value of 202409 whihc is yyyymm.


If the user's query is about new pensioners (e.g., pensioners who started receiving payments in a given month but were not present in the previous month), return a dictionary with the key 'newpensioners' and the value as the month number like if august it should be 8. If the month is not specified, ask the user for clarification.

nwr_master_data: Stores master pensioner data, including ppo_number, name, dob, pension_start_date, account_number, and age.
nwr_zone_data: Maps pensioners to zones with details like ppo_zone_code, pensioner_id, old_ppo, new_ppo, emp_name, gender, cessation_date, and pension_amount,efp_amount.
If a user specifies debit scroll only use debit scroll table instead of mapping with nwr zone data or nwr_master_data.
Rules for SQL generation:


Example queries:

"How many pensioners are registered in the system?" → SELECT COUNT(*) FROM nwr_master_data;
"Show all pensioners from a specific zone." → SELECT emp_name, new_ppo FROM nwr_zone_data WHERE ppo_zone_code = 'XYZ';
"Get total pension amount for march." → SELECT SUM(pension) FROM debit_scroll WHERE pension_month = 202403;
Return only the generated SQL query without extra commentary.Remove Any Extra Formatting (e.g., ```sql )
"Show all pensioners who are more than 80 years of age." → WITH matched_pensioners AS (
    SELECT
        ds.file_number,
        COALESCE(md1.age, md2.age) AS age
    FROM debit_scroll ds
    LEFT JOIN nwr_master_data md1 ON ds.old_ppo = md1.ppo_number
    LEFT JOIN nwr_master_data md2 ON ds.new_ppo = md2.ppo_number
    WHERE ds.pension_month = '202409'
    AND COALESCE(md1.age, md2.age) IS NOT NULL
    GROUP BY ds.file_number, age
)
SELECT COUNT(*) AS total_80_85_pensioners
FROM matched_pensioners
WHERE age >= 80 ;
"tell me total Enhance Family Pension users and normal pension users" → SELECT 
        SUM(CASE WHEN efp_amount > 0 THEN 1 ELSE 0 END) AS count_efp_greater_than_0,
        SUM(CASE WHEN efp_amount = 0 THEN 1 ELSE 0 END) AS count_efp_equal_to_0
    FROM nwr_zone_data; 
"Show me new pensioners for July 2024." → {"newpensioners": 7)}




if the user query has anything to do with new pensioners 
Only generate SELECT queries.
Allow aggregation (COUNT, SUM, AVG, etc.).
Ensure joins and filters are optimized.
Validate table and column names against the schema before generating queries.
Do not assume missing information—ask for clarification if needed.
only return sql query no need to return anything apart from it do not even add the keyword sql before it just return the query
"""


api_keyy = os.getenv("OPENAI_API_KEY")


@login_required
@csrf_exempt
def chat_completion(request):
    data = json.loads(request.body)
    user_query = data.get("query")
    client = openai.OpenAI(api_key=api_keyy)
    response = client.chat.completions.create(
    model="gpt-4o",
    messages=[
        {"role": "system", "content": db_prompt},  
        {"role": "user", "content": user_query}  
    ]
)
    new_pensioner = 0
    sql_query = response.choices[0].message.content.strip()
   
    try:
    
        sql_query_dict = json.loads(sql_query)  # Convert string to dictionary
       

        if isinstance(sql_query_dict, dict) and "newpensioners" in sql_query_dict:
            new_pensioner = 1
            func = sql_query_dict["newpensioners"]
            result = get_pension_stats_last_6_months(func)
            
        else:
            print("Key 'newpensioners' not found in the response.")
    except json.JSONDecodeError:
        pass
    

    if not is_safe_sql(sql_query):
             return JsonResponse({"error": "Invalid query: Only read operations are allowed"}, status=400)
    sql_error = None
    forward_prompt = f"""You are a **Review Bot** that analyzes user queries and SQL execution results. Your goal is to:
 Identify if the **user query** is related to pension. 
   - If **not**, politely inform the user that only pension-related queries are supported.  

 Check if the **SQL query execution** failed.  
   - If the SQL query is incorrect, analyze the database schema and provide user with a better prompt which may help the user.  
   - If the query contains invalid fields, suggest alternative fields.  

---

### **Review Process**  
 **User Query:** _"{user_query}"_  
 **Executed SQL Query:** _"{sql_query}"_  
 **SQL Error (if any):** _"{sql_error}"_  

### **Response Format**  
 If the user query is valid, respond with:  
  "Here is the requested pension information..."**  
 If the query is NOT about pension, respond politely:  
  **"I'm here to assist with pension-related queries only..."** 
    if the user query is a greeting like hey hi hello reply with greeting only. You should not never use the word sql or database in your response.
    if there is a general query like what is pension you can provide the definition of pension or any other general information about pension.
    if the user query was about active pensioner or stopped pensioner of a given month please use this data instead of the above
    "Give me the count of stopped pensioners for given month  
where inactive = stop pensioners
    if month is may -> Total Pensioners 50,579 , Active Pensioners 50,571 , Inactive Pensioners 8,Active Pension Amount ₹3,568,393,Inactive Pension Amount ₹150,973
    if month is june -> Total Pensioners 50,429 , Active Pensioners 50,355 , Inactive Pensioners 74,Active Pension Amount ₹944,578,550,Inactive Pension Amount ₹1,216,532
    if month is july -> Total Pensioners 50,523 , Active Pensioners 50,428 , Inactive Pensioners 95,Active Pension Amount ₹947,156,416,Inactive Pension Amount ₹1,349,985
    if month is august -> Total Pensioners 50,540 , Active Pensioners 50,468 , Inactive Pensioners 72,Active Pension Amount ₹951,371,149,Inactive Pension Amount ₹1,414,757
    if month is september -> Total Pensioners 50,604 , Active Pensioners 50,413 , Inactive Pensioners 191,Active Pension Amount ₹951,910,447,Inactive Pension Amount ₹3,180,245

"""
    failed_query = 0
    if new_pensioner == 0:
        try:
                with connection.cursor() as cursor:
                    cursor.execute(sql_query)
                    columns = [col[0] for col in cursor.description]
                    rows = cursor.fetchall()
                    result = [dict(zip(columns, row)) for row in rows]
        except Exception as e:
                sql_error = str(e)
                response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "system", "content": forward_prompt}]
                )
                result = response.choices[0].message.content.strip()    
                failed_query = 1
                
        if failed_query==0:
            result = [
            {columns[i]: (float(value) if isinstance(value, Decimal) else value) for i, value in enumerate(row)}
            for row in rows
            ]
    format_prompt = f"""You are a formatting expert. Please format this data based on the user query. It should be easy to read and very presentable. Also tell some basic insights about the data. Just basic insights. Use the user query to tell user from where the data is being fetched. For new pensioners the data is being fetched from debit scroll. Dont use the word database or table. If data is being fetched from nwr_zone_table just say that data is being fetched from nwr_zone sheet or record. Do this for all the data that is being fetched. 
    data: {result}
    user query: "{user_query}"
    sql query: "{sql_query}"
    """
    if failed_query==0:
        if len(result) < 20:
            response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "system", "content": format_prompt}]
            )
            result = response.choices[0].message.content.strip()        
        
    # if len(result) > 20:
#     #     result = result[:20]
#     chart_prompt = f"""
#         Generate Chart.js code for the following data:
#         {json.dumps(result)}
#         The chart should match the user's query intent: "{user_query}"
#         create only  pie chart code only with data from result. 
#         Return only JavaScript code.
#         """
#     response = client.chat.completions.create(
#     model="gpt-3.5-turbo",
#     messages=[{"role": "system", "content": chart_prompt}]
# )

  
#     chart_code = response.choices[0].message.content.strip()\
   


    return JsonResponse({"data": result,} )

@login_required
def generate_excl(request):
    rule_json = request.GET.get('rule_data')
    month = request.GET.get('month')
    year = 2024 
    month = str(month).zfill(2)
    prev_year = str(year)

    if month == "01":
        prev_month = "12"
        prev_year = str(int(year) - 1)


    curr_pension_month = f"{year}{month}"


    print(curr_pension_month)

    
    try:
        rule_data = json.loads(rule_json)
    except json.JSONDecodeError:
        return HttpResponse("Invalid JSON", status=400)

    
    wb = openpyxl.Workbook()
    ws = wb.active
    print(rule_data,"rule_data")
    
    rule = 2  
    if isinstance(rule_data, dict):
        rule_data = [rule_data]

    if not isinstance(rule_data, list) or not rule_data or not isinstance(rule_data[0], dict):
        return HttpResponse("Invalid data format", status=400)

    if 'matched_records' in rule_json:
        rule = 1
    elif 'regular_pension_count' in rule_json:
        rule = 3
    elif 'stats' in rule_json:
        rule = 5
    elif 'new' in rule_json:
        rule = 4

    if rule == 1:
        headers = list(rule_data[0].keys())

        ws.title = "Exported Data"

        ws.append(headers)

        for item in rule_data:
            row = [item.get(col, "") for col in headers]
            ws.append(row)

        unlink_sheet = wb.create_sheet(title="UNLINKED RECORDS")
        underpayment_sheet = wb.create_sheet(title="UNDERPAYMENT RECORDS")
        overpayment_sheet = wb.create_sheet(title="OVERPAYMENT RECORDS")
        
        query = """
           SELECT e.*, md.* 
            FROM arpan_exception AS e 
            LEFT JOIN nwr_master_data AS md 
            ON md.ppo_number = e.scroll_ppo_no
            WHERE e.ifsc_code LIKE '%SBI%';
        """
        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            results = [dict(zip(columns, row)) for row in cursor.fetchall()]

        unlink_df = pd.DataFrame(results)
        unlink_df_temp = unlink_df.fillna('')
        unlink_df = unlink_df_temp.drop("id",axis = 1)
        # Underpayment Query
        query_underpayment = """
            SELECT e.*, md.* 
            FROM mismatch_data AS e  
            LEFT JOIN nwr_master_data AS md 
            ON md.ppo_number = e.ppo_number 
            WHERE e.basic_diff < 0 AND e.month = %s
        """
        with connection.cursor() as cursor:
            cursor.execute(query_underpayment, [curr_pension_month])
            columns = [col[0] for col in cursor.description]
            underpayment_results = [dict(zip(columns, row)) for row in cursor.fetchall()]

        underpayment_results_df = pd.DataFrame(underpayment_results)
        underpayment_results_df_temp = underpayment_results_df.drop("id", axis= 1)
        underpayment_results_df = underpayment_results_df_temp.fillna('')
       

        
        query_overpayment = """
            SELECT e.*, md.* 
            FROM mismatch_data AS e  
            LEFT JOIN nwr_master_data AS md 
            ON md.ppo_number = e.ppo_number 
            WHERE e.basic_diff > 0 AND e.month = %s
        """
        with connection.cursor() as cursor:
            cursor.execute(query_overpayment, [curr_pension_month])
            columns = [col[0] for col in cursor.description]
            overpayment_results = [dict(zip(columns, row)) for row in cursor.fetchall()]

        
        overpayment_results_df_temp = pd.DataFrame(overpayment_results)
        overpayment_results_df =  overpayment_results_df_temp.drop("id",axis = 1)
        
        overpayment_results_df = overpayment_results_df.fillna('')

        insert_df_into_sheet(unlink_df, unlink_sheet)
        insert_df_into_sheet(underpayment_results_df, underpayment_sheet)
        insert_df_into_sheet(overpayment_results_df, overpayment_sheet)

    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    response = HttpResponse(
        excel_stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="exported_data.xlsx"'
    return response

def insert_df_into_sheet(df, sheet):
    if not df.empty:
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=value)


# def generate_data_excel():
    # Fetch all required data in one query
    # pension_months = ["202409", "202408", "202407", "202406", "202405"]
    # debit_scrolls = DebitScroll.objects.filter(pension_month__in=pension_months).values()

    # # Convert QuerySet to DataFrame
    # df = pd.DataFrame(list(debit_scrolls))

    # if df.empty:
    #     return "No data available"

    # # Group by 'pension_month'
    # grouped_df = df.groupby("pension_month")

    # # Define the Excel file path
    # file_path = "debit_scrolls.xlsx"

    # # Write to Excel with each pension_month as a separate sheet
    # with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
    #     for pension_month, group in grouped_df:
    #         sheet_name = f"{pension_month}_Data"
    #         group.to_excel(writer, sheet_name=sheet_name, index=False)
    # zone_data = nwr_zone_data.objects.all().values()  
    # # Convert to DataFrame
    # zone_data_df = pd.DataFrame(list(zone_data))
    # # Save to Excel
    # file_path = "nwr_zone_data.xlsx"
    # with pd.ExcelWriter(file_path, engine="xlsxwriter") as writer:
    #     zone_data_df.to_excel(writer, sheet_name="Sheet1", index=False)

    # master_data = NWRMasterData.objects.all().values()
    # master_data_df = pd.DataFrame(list(master_data))
    # file_path = "Master_data_sheet.xlsx"
    # with pd.ExcelWriter(file_path,engine = "xlsxwriter") as writer:
    #     master_data_df.to_excel(writer, sheet_name="Sheet1", index=False)



    
    # commutation_data = mismatch_data.objects.all().values()
    # commutation_data_df = pd.DataFrame(list(commutation_data))
    # file_path = "mismatch_data.xlsx"
    # with pd.ExcelWriter(file_path,engine = "xlsxwriter") as writer:
    #    commutation_data_df.to_excel(writer, sheet_name="Sheet1", index=False)
    # print("function_worked")